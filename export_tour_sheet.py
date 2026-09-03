#!/usr/bin/env python3
"""v_tour_sheet_rows 를 '글로벌 투어 현황' 시트 양식의 xlsx 로 내보낸다.

기존 시트에 바로 붙여넣을 수 있도록 컬럼 순서와 헤더 문구를 그대로 맞췄다.

아직 못 채우는 칸은 회색으로 칠하고 비워둔다 (공연 규모, 판매 완료 좌석수,
Regular/VIP 티켓 가격). 트윗 본문에 없는 정보라 판매 링크를 타고 들어가거나
베뉴 마스터가 있어야 나온다. 추정치로 채우면 사람이 검증을 안 하게 되므로
빈칸이 낫다.

환경변수
  TOUR_SHEET_OUT : 출력 파일 경로 (기본 글로벌투어_자동수집_<오늘>.xlsx)
"""
import datetime
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bq_common import DATASET, PROJECT_ID, get_bq_client

VIEW = f"{PROJECT_ID}.{DATASET}.v_tour_sheet_rows"

HEADERS = [
    "공연 일자", "공연 유형", "단독 이벤트 제안 여부\n(미주유럽사업팀 내부 참고용)",
    "공연명", "IP", "리전", "국가", "도시", "베뉴명",
    "공연 규모 \n(판매 좌석수)", "판매 완료 좌석 수",
    "Regular 티켓 가격", "VIP 티켓 가격", "판매 링크", "작성 일자", "비고",
]
WIDTHS = [12, 14, 16, 42, 18, 10, 16, 18, 34, 12, 12, 14, 14, 40, 12, 62]
# 아직 못 채우는 칸 (1-based 컬럼 번호)
GAP_COLS = (10, 11, 12, 13)
ARIAL = "Arial"


def fetch(bq):
    return list(bq.query(f"""
        SELECT `공연_일자`, `공연_유형`, `단독_이벤트_제안_여부`, `공연명`, `IP`,
               `리전`, `국가`, `도시`, `베뉴명`, `판매_링크`, `작성_일자`, `비고`,
               needs_review
        FROM `{VIEW}`
        ORDER BY `공연_일자`, `IP`, `도시`
    """).result())


def build(rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "글로벌 투어 현황_자동수집"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    gap_fill = PatternFill("solid", fgColor="F2F2F2")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")

    ws["A1"] = "X 공식계정 자동 수집 결과 — '글로벌 투어 현황' 시트 양식"
    ws["A1"].font = Font(name=ARIAL, size=13, bold=True)
    ws["A2"] = (
        f"생성일 {datetime.date.today():%Y-%m-%d} · {len(rows)}행 · 출처 BigQuery {VIEW}\n"
        "회색 칸(공연 규모·판매 완료 좌석수·티켓 가격)은 트윗 본문에 없는 정보다. "
        "판매 링크를 타고 들어가거나 베뉴 마스터가 있어야 채울 수 있다. 빈칸으로 두었다.\n"
        "주황색 행은 자동 판정이 확신하지 못한 건이다. 비고의 원문 링크로 확인이 필요하다.\n"
        "미국 도시는 시트 규칙상 주(州)까지 적어야 하는데, 트윗에 주 표기가 없어 도시명만 들어갔다."
    )
    ws["A2"].font = Font(name=ARIAL, size=9)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].fill = note_fill
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.row_dimensions[2].height = 62

    hdr_row = 4
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(hdr_row, c, h)
        cell.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    ws.row_dimensions[hdr_row].height = 34

    for i, r in enumerate(rows):
        row = hdr_row + 1 + i
        vals = [
            r["공연_일자"], r["공연_유형"], r["단독_이벤트_제안_여부"], r["공연명"], r["IP"],
            r["리전"], r["국가"], r["도시"], r["베뉴명"],
            None, None, None, None,
            r["판매_링크"], r["작성_일자"], r["비고"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, v)
            cell.font = Font(name=ARIAL, size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (4, 16)))
            if isinstance(v, datetime.date):
                cell.number_format = "yyyy-mm-dd"
            if c in GAP_COLS:
                cell.fill = gap_fill
            elif r["needs_review"]:
                cell.fill = warn_fill
        if r["판매_링크"]:
            link = ws.cell(row, 14)
            link.hyperlink = r["판매_링크"]
            link.font = Font(name=ARIAL, size=9, color="0563C1", underline="single")

    for c, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    last = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A{hdr_row}:{last}{hdr_row + len(rows)}"
    wb.save(out_path)


def main():
    bq = get_bq_client()
    rows = fetch(bq)
    if not rows:
        print("내보낼 행이 없다")
        return
    out = os.environ.get(
        "TOUR_SHEET_OUT",
        f"글로벌투어_자동수집_{datetime.date.today():%Y%m%d}.xlsx")
    build(rows, out)
    review = sum(1 for r in rows if r["needs_review"])
    venue = sum(1 for r in rows if r["베뉴명"])
    print(f"{out} 저장 | {len(rows)}행 (확인 필요 {review}, 베뉴 확보 {venue})")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()
        sys.exit(1)
